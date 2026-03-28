"""
Script unificado para gerenciamento de faturas
"""
import sqlite3
from datetime import datetime, date, timedelta
import sys
import os

def limpar_tela():
    """Limpa a tela do terminal"""
    os.system('cls' if os.name == 'nt' else 'clear')

def gerar_relatorio_unico(data_execucao):
    """
    Gera relatório XLSX para uma data específica
    
    Args:
        data_execucao (str): Data no formato YYYY-MM-DD
    
    Returns:
        str: Nome do arquivo gerado ou None
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("❌ Biblioteca openpyxl não encontrada!")
        print("📦 Instalando openpyxl...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    
    # Conectar ao banco
    conn = sqlite3.connect('database/faturas.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Buscar faturas processadas no dia
    cursor.execute("""
        SELECT 
            id, nova_uc, mes_referencia, cnpj_geradora, 
            status, tipo_operacao, valor, data_vencimento, 
            situacao_pagamento, data_processamento, tentativas, 
            mensagem_erro
        FROM faturas
        WHERE DATE(data_processamento) = ?
        ORDER BY cnpj_geradora, nova_uc, mes_referencia
    """, (data_execucao,))
    
    faturas = [dict(row) for row in cursor.fetchall()]
    
    # Buscar execuções diárias
    cursor.execute("""
        SELECT 
            cnpj_geradora, nova_uc, total_faturas, 
            faturas_sucesso, faturas_erro, faturas_puladas,
            status_execucao, data_hora_inicio, data_hora_fim
        FROM execucoes_diarias
        WHERE data_execucao = ?
        ORDER BY cnpj_geradora, nova_uc
    """, (data_execucao,))
    
    execucoes = [dict(row) for row in cursor.fetchall()]
    
    conn.close()
    
    if not faturas and not execucoes:
        return None
    
    # Criar workbook
    wb = Workbook()
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ==================== ABA 1: RESUMO ====================
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    
    ws_resumo['A1'] = f"RELATÓRIO DE EXECUÇÃO - {datetime.strptime(data_execucao, '%Y-%m-%d').strftime('%d/%m/%Y')}"
    ws_resumo['A1'].font = Font(bold=True, size=14)
    ws_resumo.merge_cells('A1:F1')
    
    total_faturas = len(faturas)
    total_sucesso = sum(1 for f in faturas if f['status'] == 'sucesso')
    total_erro = sum(1 for f in faturas if f['status'] == 'erro')
    
    ws_resumo['A3'] = "ESTATÍSTICAS GERAIS"
    ws_resumo['A3'].font = Font(bold=True, size=12)
    
    ws_resumo['A4'] = "Total de Faturas Processadas:"
    ws_resumo['B4'] = total_faturas
    ws_resumo['A5'] = "Sucesso:"
    ws_resumo['B5'] = total_sucesso
    ws_resumo['A6'] = "Erro:"
    ws_resumo['B6'] = total_erro
    ws_resumo['A7'] = "Taxa de Sucesso:"
    ws_resumo['B7'] = f"{(total_sucesso/total_faturas*100):.1f}%" if total_faturas > 0 else "0%"
    
    ws_resumo['A9'] = "POR TIPO DE OPERAÇÃO"
    ws_resumo['A9'].font = Font(bold=True, size=12)
    
    tipos_operacao = {}
    for f in faturas:
        tipo = f['tipo_operacao'] or 'Não registrado'
        tipos_operacao[tipo] = tipos_operacao.get(tipo, 0) + 1
    
    linha = 10
    for tipo, count in sorted(tipos_operacao.items()):
        ws_resumo[f'A{linha}'] = f"{tipo}:"
        ws_resumo[f'B{linha}'] = count
        linha += 1
    
    ws_resumo[f'A{linha+1}'] = "POR GERADORA"
    ws_resumo[f'A{linha+1}'].font = Font(bold=True, size=12)
    
    geradoras = {}
    for f in faturas:
        cnpj = f['cnpj_geradora']
        if cnpj not in geradoras:
            geradoras[cnpj] = {'total': 0, 'sucesso': 0, 'erro': 0}
        geradoras[cnpj]['total'] += 1
        if f['status'] == 'sucesso':
            geradoras[cnpj]['sucesso'] += 1
        else:
            geradoras[cnpj]['erro'] += 1
    
    linha += 3
    ws_resumo[f'A{linha}'] = "CNPJ"
    ws_resumo[f'B{linha}'] = "Total"
    ws_resumo[f'C{linha}'] = "Sucesso"
    ws_resumo[f'D{linha}'] = "Erro"
    ws_resumo[f'E{linha}'] = "Taxa"
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_resumo[f'{col}{linha}'].fill = header_fill
        ws_resumo[f'{col}{linha}'].font = header_font
        ws_resumo[f'{col}{linha}'].border = border
    
    linha += 1
    for cnpj, stats in sorted(geradoras.items()):
        ws_resumo[f'A{linha}'] = cnpj
        ws_resumo[f'B{linha}'] = stats['total']
        ws_resumo[f'C{linha}'] = stats['sucesso']
        ws_resumo[f'D{linha}'] = stats['erro']
        ws_resumo[f'E{linha}'] = f"{(stats['sucesso']/stats['total']*100):.1f}%"
        linha += 1
    
    ws_resumo.column_dimensions['A'].width = 30
    ws_resumo.column_dimensions['B'].width = 15
    ws_resumo.column_dimensions['C'].width = 15
    ws_resumo.column_dimensions['D'].width = 15
    ws_resumo.column_dimensions['E'].width = 15
    
    # ==================== ABA 2: FATURAS ====================
    ws_faturas = wb.create_sheet("Faturas Processadas")
    
    headers = [
        "ID", "UC", "Mês Ref", "CNPJ", "Status", "Tipo Operação",
        "Valor", "Vencimento", "Situação Pgto", "Data Processamento",
        "Tentativas", "Mensagem Erro"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_faturas.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for row_num, fatura in enumerate(faturas, 2):
        ws_faturas.cell(row=row_num, column=1, value=fatura['id'])
        ws_faturas.cell(row=row_num, column=2, value=fatura['nova_uc'])
        ws_faturas.cell(row=row_num, column=3, value=fatura['mes_referencia'])
        ws_faturas.cell(row=row_num, column=4, value=fatura['cnpj_geradora'])
        ws_faturas.cell(row=row_num, column=5, value=fatura['status'])
        ws_faturas.cell(row=row_num, column=6, value=fatura['tipo_operacao'] or 'N/A')
        ws_faturas.cell(row=row_num, column=7, value=f"R$ {fatura['valor']}" if fatura['valor'] else 'N/A')
        ws_faturas.cell(row=row_num, column=8, value=fatura['data_vencimento'] or 'N/A')
        ws_faturas.cell(row=row_num, column=9, value=fatura['situacao_pagamento'] or 'N/A')
        ws_faturas.cell(row=row_num, column=10, value=fatura['data_processamento'])
        ws_faturas.cell(row=row_num, column=11, value=fatura['tentativas'])
        ws_faturas.cell(row=row_num, column=12, value=fatura['mensagem_erro'] or '')
        
        fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") if fatura['status'] == 'sucesso' else PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        for col_num in range(1, 13):
            ws_faturas.cell(row=row_num, column=col_num).fill = fill
            ws_faturas.cell(row=row_num, column=col_num).border = border
    
    ws_faturas.column_dimensions['A'].width = 10
    ws_faturas.column_dimensions['B'].width = 15
    ws_faturas.column_dimensions['C'].width = 12
    ws_faturas.column_dimensions['D'].width = 20
    ws_faturas.column_dimensions['E'].width = 12
    ws_faturas.column_dimensions['F'].width = 18
    ws_faturas.column_dimensions['G'].width = 15
    ws_faturas.column_dimensions['H'].width = 15
    ws_faturas.column_dimensions['I'].width = 15
    ws_faturas.column_dimensions['J'].width = 20
    ws_faturas.column_dimensions['K'].width = 12
    ws_faturas.column_dimensions['L'].width = 40
    
    # ==================== ABA 3: EXECUÇÕES ====================
    ws_execucoes = wb.create_sheet("Execuções por UC")
    
    headers_exec = [
        "CNPJ", "UC", "Total", "Sucesso", "Erro", 
        "Puladas", "Status", "Hora Início", "Hora Fim", "Duração"
    ]
    
    for col_num, header in enumerate(headers_exec, 1):
        cell = ws_execucoes.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for row_num, exec in enumerate(execucoes, 2):
        ws_execucoes.cell(row=row_num, column=1, value=exec['cnpj_geradora'])
        ws_execucoes.cell(row=row_num, column=2, value=exec['nova_uc'])
        ws_execucoes.cell(row=row_num, column=3, value=exec['total_faturas'])
        ws_execucoes.cell(row=row_num, column=4, value=exec['faturas_sucesso'])
        ws_execucoes.cell(row=row_num, column=5, value=exec['faturas_erro'])
        ws_execucoes.cell(row=row_num, column=6, value=exec['faturas_puladas'])
        ws_execucoes.cell(row=row_num, column=7, value=exec['status_execucao'])
        ws_execucoes.cell(row=row_num, column=8, value=exec['data_hora_inicio'])
        ws_execucoes.cell(row=row_num, column=9, value=exec['data_hora_fim'] or 'N/A')
        
        if exec['data_hora_fim']:
            try:
                inicio = datetime.strptime(exec['data_hora_inicio'], '%Y-%m-%d %H:%M:%S.%f')
                fim = datetime.strptime(exec['data_hora_fim'], '%Y-%m-%d %H:%M:%S.%f')
                duracao = fim - inicio
                ws_execucoes.cell(row=row_num, column=10, value=str(duracao).split('.')[0])
            except:
                ws_execucoes.cell(row=row_num, column=10, value='N/A')
        else:
            ws_execucoes.cell(row=row_num, column=10, value='N/A')
        
        if exec['status_execucao'] == 'completo':
            fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif exec['status_execucao'] == 'parcial':
            fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        for col_num in range(1, 11):
            ws_execucoes.cell(row=row_num, column=col_num).fill = fill
            ws_execucoes.cell(row=row_num, column=col_num).border = border
    
    ws_execucoes.column_dimensions['A'].width = 20
    ws_execucoes.column_dimensions['B'].width = 15
    ws_execucoes.column_dimensions['C'].width = 12
    ws_execucoes.column_dimensions['D'].width = 12
    ws_execucoes.column_dimensions['E'].width = 12
    ws_execucoes.column_dimensions['F'].width = 12
    ws_execucoes.column_dimensions['G'].width = 15
    ws_execucoes.column_dimensions['H'].width = 20
    ws_execucoes.column_dimensions['I'].width = 20
    ws_execucoes.column_dimensions['J'].width = 15
    
    # Salvar arquivo
    data_formatada = datetime.strptime(data_execucao, '%Y-%m-%d').strftime('%d%m%Y')
    nome_arquivo = f"relatorio_faturas_{data_formatada}.xlsx"
    
    wb.save(nome_arquivo)
    
    print(f"\n✅ Relatório gerado com sucesso: {nome_arquivo}")
    print(f"📊 Total de faturas: {total_faturas}")
    print(f"✅ Sucesso: {total_sucesso}")
    print(f"❌ Erro: {total_erro}")
    print(f"📈 Taxa de sucesso: {(total_sucesso/total_faturas*100):.1f}%" if total_faturas > 0 else "0%")
    
    return nome_arquivo

def gerar_relatorio_hoje():
    """Gera relatório do dia atual"""
    data_hoje = date.today().strftime("%Y-%m-%d")
    print(f"\n📅 Gerando relatório para hoje: {data_hoje}")
    
    arquivo = gerar_relatorio_unico(data_hoje)
    
    if arquivo:
        abrir = input("\nDeseja abrir o arquivo agora? (s/n): ").strip().lower()
        if abrir == 's':
            try:
                os.startfile(arquivo)
                print("📂 Abrindo arquivo...")
            except:
                print("⚠️ Não foi possível abrir automaticamente. Abra manualmente.")
        
        input("\nPressione ENTER para continuar...")
    else:
        print("\n⚠️ Nenhum dado encontrado para hoje")
        input("\nPressione ENTER para continuar...")

def gerar_relatorio_intervalo():
    """Gera relatório com intervalo de datas"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("❌ Biblioteca openpyxl não encontrada!")
        print("📦 Instalando openpyxl...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    
    print("\n" + "=" * 80)
    print("RELATÓRIO COM INTERVALO DE DATAS")
    print("=" * 80)
    print("\nFormato da data: DD/MM/YYYY")
    print("Exemplo: 27/03/2026")
    print()
    
    # Solicitar data inicial
    while True:
        data_inicial_str = input("Data inicial: ").strip()
        try:
            data_inicial = datetime.strptime(data_inicial_str, "%d/%m/%Y").date()
            break
        except ValueError:
            print("❌ Formato inválido! Use DD/MM/YYYY")
    
    # Solicitar data final
    while True:
        data_final_str = input("Data final: ").strip()
        try:
            data_final = datetime.strptime(data_final_str, "%d/%m/%Y").date()
            if data_final < data_inicial:
                print("❌ Data final deve ser maior ou igual à data inicial!")
                continue
            break
        except ValueError:
            print("❌ Formato inválido! Use DD/MM/YYYY")
    
    print(f"\n📊 Buscando dados de {data_inicial_str} até {data_final_str}...")
    
    # Conectar ao banco
    conn = sqlite3.connect('database/faturas.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Buscar faturas do período
    cursor.execute("""
        SELECT 
            id, nova_uc, mes_referencia, cnpj_geradora, 
            status, tipo_operacao, valor, data_vencimento, 
            situacao_pagamento, data_processamento, tentativas, 
            mensagem_erro
        FROM faturas
        WHERE DATE(data_processamento) BETWEEN ? AND ?
        ORDER BY data_processamento, cnpj_geradora, nova_uc
    """, (data_inicial.strftime("%Y-%m-%d"), data_final.strftime("%Y-%m-%d")))
    
    faturas = [dict(row) for row in cursor.fetchall()]
    
    # Buscar execuções do período
    cursor.execute("""
        SELECT 
            data_execucao, cnpj_geradora, nova_uc, total_faturas, 
            faturas_sucesso, faturas_erro, faturas_puladas,
            status_execucao, data_hora_inicio, data_hora_fim
        FROM execucoes_diarias
        WHERE data_execucao BETWEEN ? AND ?
        ORDER BY data_execucao, cnpj_geradora, nova_uc
    """, (data_inicial.strftime("%Y-%m-%d"), data_final.strftime("%Y-%m-%d")))
    
    execucoes = [dict(row) for row in cursor.fetchall()]
    
    conn.close()
    
    if not faturas and not execucoes:
        print(f"\n⚠️ Nenhum dado encontrado para o período")
        input("\nPressione ENTER para continuar...")
        return
    
    # Criar workbook
    wb = Workbook()
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ==================== ABA 1: RESUMO ====================
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    
    # Título
    ws_resumo['A1'] = f"RELATÓRIO DE EXECUÇÃO - {data_inicial_str} a {data_final_str}"
    ws_resumo['A1'].font = Font(bold=True, size=14)
    ws_resumo.merge_cells('A1:F1')
    
    # Estatísticas gerais
    total_faturas = len(faturas)
    total_sucesso = sum(1 for f in faturas if f['status'] == 'sucesso')
    total_erro = sum(1 for f in faturas if f['status'] == 'erro')
    
    ws_resumo['A3'] = "ESTATÍSTICAS GERAIS"
    ws_resumo['A3'].font = Font(bold=True, size=12)
    
    ws_resumo['A4'] = "Total de Faturas Processadas:"
    ws_resumo['B4'] = total_faturas
    ws_resumo['A5'] = "Sucesso:"
    ws_resumo['B5'] = total_sucesso
    ws_resumo['A6'] = "Erro:"
    ws_resumo['B6'] = total_erro
    ws_resumo['A7'] = "Taxa de Sucesso:"
    ws_resumo['B7'] = f"{(total_sucesso/total_faturas*100):.1f}%" if total_faturas > 0 else "0%"
    
    # Por tipo de operação
    ws_resumo['A9'] = "POR TIPO DE OPERAÇÃO"
    ws_resumo['A9'].font = Font(bold=True, size=12)
    
    tipos_operacao = {}
    for f in faturas:
        tipo = f['tipo_operacao'] or 'Não registrado'
        tipos_operacao[tipo] = tipos_operacao.get(tipo, 0) + 1
    
    linha = 10
    for tipo, count in sorted(tipos_operacao.items()):
        ws_resumo[f'A{linha}'] = f"{tipo}:"
        ws_resumo[f'B{linha}'] = count
        linha += 1
    
    # Por geradora
    ws_resumo[f'A{linha+1}'] = "POR GERADORA"
    ws_resumo[f'A{linha+1}'].font = Font(bold=True, size=12)
    
    geradoras = {}
    for f in faturas:
        cnpj = f['cnpj_geradora']
        if cnpj not in geradoras:
            geradoras[cnpj] = {'total': 0, 'sucesso': 0, 'erro': 0}
        geradoras[cnpj]['total'] += 1
        if f['status'] == 'sucesso':
            geradoras[cnpj]['sucesso'] += 1
        else:
            geradoras[cnpj]['erro'] += 1
    
    linha += 3
    ws_resumo[f'A{linha}'] = "CNPJ"
    ws_resumo[f'B{linha}'] = "Total"
    ws_resumo[f'C{linha}'] = "Sucesso"
    ws_resumo[f'D{linha}'] = "Erro"
    ws_resumo[f'E{linha}'] = "Taxa"
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_resumo[f'{col}{linha}'].fill = header_fill
        ws_resumo[f'{col}{linha}'].font = header_font
        ws_resumo[f'{col}{linha}'].border = border
    
    linha += 1
    for cnpj, stats in sorted(geradoras.items()):
        ws_resumo[f'A{linha}'] = cnpj
        ws_resumo[f'B{linha}'] = stats['total']
        ws_resumo[f'C{linha}'] = stats['sucesso']
        ws_resumo[f'D{linha}'] = stats['erro']
        ws_resumo[f'E{linha}'] = f"{(stats['sucesso']/stats['total']*100):.1f}%"
        linha += 1
    
    ws_resumo.column_dimensions['A'].width = 30
    ws_resumo.column_dimensions['B'].width = 15
    ws_resumo.column_dimensions['C'].width = 15
    ws_resumo.column_dimensions['D'].width = 15
    ws_resumo.column_dimensions['E'].width = 15
    
    # ==================== ABA 2: FATURAS ====================
    ws_faturas = wb.create_sheet("Faturas Processadas")
    
    headers = [
        "ID", "UC", "Mês Ref", "CNPJ", "Status", "Tipo Operação",
        "Valor", "Vencimento", "Situação Pgto", "Data Processamento",
        "Tentativas", "Mensagem Erro"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_faturas.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for row_num, fatura in enumerate(faturas, 2):
        ws_faturas.cell(row=row_num, column=1, value=fatura['id'])
        ws_faturas.cell(row=row_num, column=2, value=fatura['nova_uc'])
        ws_faturas.cell(row=row_num, column=3, value=fatura['mes_referencia'])
        ws_faturas.cell(row=row_num, column=4, value=fatura['cnpj_geradora'])
        ws_faturas.cell(row=row_num, column=5, value=fatura['status'])
        ws_faturas.cell(row=row_num, column=6, value=fatura['tipo_operacao'] or 'N/A')
        ws_faturas.cell(row=row_num, column=7, value=f"R$ {fatura['valor']}" if fatura['valor'] else 'N/A')
        ws_faturas.cell(row=row_num, column=8, value=fatura['data_vencimento'] or 'N/A')
        ws_faturas.cell(row=row_num, column=9, value=fatura['situacao_pagamento'] or 'N/A')
        ws_faturas.cell(row=row_num, column=10, value=fatura['data_processamento'])
        ws_faturas.cell(row=row_num, column=11, value=fatura['tentativas'])
        ws_faturas.cell(row=row_num, column=12, value=fatura['mensagem_erro'] or '')
        
        fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") if fatura['status'] == 'sucesso' else PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        for col_num in range(1, 13):
            ws_faturas.cell(row=row_num, column=col_num).fill = fill
            ws_faturas.cell(row=row_num, column=col_num).border = border
    
    ws_faturas.column_dimensions['A'].width = 10
    ws_faturas.column_dimensions['B'].width = 15
    ws_faturas.column_dimensions['C'].width = 12
    ws_faturas.column_dimensions['D'].width = 20
    ws_faturas.column_dimensions['E'].width = 12
    ws_faturas.column_dimensions['F'].width = 18
    ws_faturas.column_dimensions['G'].width = 15
    ws_faturas.column_dimensions['H'].width = 15
    ws_faturas.column_dimensions['I'].width = 15
    ws_faturas.column_dimensions['J'].width = 20
    ws_faturas.column_dimensions['K'].width = 12
    ws_faturas.column_dimensions['L'].width = 40
    
    # ==================== ABA 3: EXECUÇÕES ====================
    ws_execucoes = wb.create_sheet("Execuções por UC")
    
    headers_exec = [
        "Data", "CNPJ", "UC", "Total", "Sucesso", "Erro", 
        "Puladas", "Status", "Hora Início", "Hora Fim", "Duração"
    ]
    
    for col_num, header in enumerate(headers_exec, 1):
        cell = ws_execucoes.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for row_num, exec in enumerate(execucoes, 2):
        ws_execucoes.cell(row=row_num, column=1, value=exec['data_execucao'])
        ws_execucoes.cell(row=row_num, column=2, value=exec['cnpj_geradora'])
        ws_execucoes.cell(row=row_num, column=3, value=exec['nova_uc'])
        ws_execucoes.cell(row=row_num, column=4, value=exec['total_faturas'])
        ws_execucoes.cell(row=row_num, column=5, value=exec['faturas_sucesso'])
        ws_execucoes.cell(row=row_num, column=6, value=exec['faturas_erro'])
        ws_execucoes.cell(row=row_num, column=7, value=exec['faturas_puladas'])
        ws_execucoes.cell(row=row_num, column=8, value=exec['status_execucao'])
        ws_execucoes.cell(row=row_num, column=9, value=exec['data_hora_inicio'])
        ws_execucoes.cell(row=row_num, column=10, value=exec['data_hora_fim'] or 'N/A')
        
        if exec['data_hora_fim']:
            try:
                inicio = datetime.strptime(exec['data_hora_inicio'], '%Y-%m-%d %H:%M:%S.%f')
                fim = datetime.strptime(exec['data_hora_fim'], '%Y-%m-%d %H:%M:%S.%f')
                duracao = fim - inicio
                ws_execucoes.cell(row=row_num, column=11, value=str(duracao).split('.')[0])
            except:
                ws_execucoes.cell(row=row_num, column=11, value='N/A')
        else:
            ws_execucoes.cell(row=row_num, column=11, value='N/A')
        
        if exec['status_execucao'] == 'completo':
            fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif exec['status_execucao'] == 'parcial':
            fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        for col_num in range(1, 12):
            ws_execucoes.cell(row=row_num, column=col_num).fill = fill
            ws_execucoes.cell(row=row_num, column=col_num).border = border
    
    ws_execucoes.column_dimensions['A'].width = 12
    ws_execucoes.column_dimensions['B'].width = 20
    ws_execucoes.column_dimensions['C'].width = 15
    ws_execucoes.column_dimensions['D'].width = 12
    ws_execucoes.column_dimensions['E'].width = 12
    ws_execucoes.column_dimensions['F'].width = 12
    ws_execucoes.column_dimensions['G'].width = 12
    ws_execucoes.column_dimensions['H'].width = 15
    ws_execucoes.column_dimensions['I'].width = 20
    ws_execucoes.column_dimensions['J'].width = 20
    ws_execucoes.column_dimensions['K'].width = 15
    
    # Salvar arquivo
    nome_arquivo = f"relatorio_faturas_{data_inicial.strftime('%d%m%Y')}_a_{data_final.strftime('%d%m%Y')}.xlsx"
    wb.save(nome_arquivo)
    
    print(f"\n✅ Relatório gerado com sucesso: {nome_arquivo}")
    print(f"📊 Total de faturas: {total_faturas}")
    print(f"✅ Sucesso: {total_sucesso}")
    print(f"❌ Erro: {total_erro}")
    print(f"📈 Taxa de sucesso: {(total_sucesso/total_faturas*100):.1f}%" if total_faturas > 0 else "0%")
    
    abrir = input("\nDeseja abrir o arquivo agora? (s/n): ").strip().lower()
    if abrir == 's':
        try:
            os.startfile(nome_arquivo)
            print("📂 Abrindo arquivo...")
        except:
            print("⚠️ Não foi possível abrir automaticamente. Abra manualmente.")
    
    input("\nPressione ENTER para continuar...")

def verificar_faturas_pendentes():
    """Verifica se existem faturas com status a_verificar"""
    print("\n" + "=" * 80)
    print("VERIFICAÇÃO DE FATURAS PENDENTES")
    print("=" * 80)
    print()
    
    conn = sqlite3.connect('database/faturas.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Contar faturas por status
    cursor.execute("""
        SELECT status, COUNT(*) as total
        FROM faturas
        GROUP BY status
    """)
    
    print("📊 DISTRIBUIÇÃO POR STATUS:")
    print("-" * 80)
    
    status_counts = {}
    for row in cursor.fetchall():
        status_counts[row['status']] = row['total']
        print(f"   {row['status']}: {row['total']} faturas")
    
    print()
    
    # Verificar faturas a_verificar
    total_pendentes = status_counts.get('a_verificar', 0)
    
    if total_pendentes > 0:
        print(f"⚠️ Existem {total_pendentes} faturas com status 'a_verificar'")
        print()
        
        # Mostrar detalhes por geradora
        cursor.execute("""
            SELECT cnpj_geradora, COUNT(*) as total
            FROM faturas
            WHERE status = 'a_verificar'
            GROUP BY cnpj_geradora
            ORDER BY total DESC
        """)
        
        print("📋 FATURAS PENDENTES POR GERADORA:")
        print("-" * 80)
        for row in cursor.fetchall():
            print(f"   {row['cnpj_geradora']}: {row['total']} faturas")
        
        print()
        
        # Perguntar se quer ver detalhes
        ver_detalhes = input("Deseja ver a lista completa de faturas pendentes? (s/n): ").strip().lower()
        
        if ver_detalhes == 's':
            cursor.execute("""
                SELECT id, nova_uc, mes_referencia, cnpj_geradora, data_criacao
                FROM faturas
                WHERE status = 'a_verificar'
                ORDER BY cnpj_geradora, nova_uc, mes_referencia
                LIMIT 50
            """)
            
            print("\n📋 PRIMEIRAS 50 FATURAS PENDENTES:")
            print("-" * 80)
            for row in cursor.fetchall():
                print(f"   ID: {row['id']} | UC: {row['nova_uc']} | Mês: {row['mes_referencia']} | CNPJ: {row['cnpj_geradora']}")
    else:
        print("✅ Não existem faturas pendentes (a_verificar)")
        print("🎉 Todas as faturas foram processadas!")
    
    # Verificar faturas com erro
    total_erros = status_counts.get('erro', 0)
    
    if total_erros > 0:
        print()
        print(f"⚠️ Existem {total_erros} faturas com status 'erro'")
        print("💡 Use 'python robo.py --force' para reprocessar faturas com erro")
    
    conn.close()
    
    input("\nPressione ENTER para continuar...")

def menu_principal():
    """Menu principal do sistema"""
    while True:
        limpar_tela()
        print("=" * 80)
        print("GERENCIAMENTO DE FATURAS - SISTEMA DE RELATÓRIOS")
        print("=" * 80)
        print("""

███████╗ █████╗     ███╗   ███╗ █████╗ ███╗   ██╗ █████╗  ██████╗ ███████╗██████╗ 
██╔════╝██╔══██╗    ████╗ ████║██╔══██╗████╗  ██║██╔══██╗██╔════╝ ██╔════╝██╔══██╗
█████╗  ███████║    ██╔████╔██║███████║██╔██╗ ██║███████║██║  ███╗█████╗  ██████╔╝
██╔══╝  ██╔══██║    ██║╚██╔╝██║██╔══██║██║╚██╗██║██╔══██║██║   ██║██╔══╝  ██╔══██╗
███████╗██║  ██║    ██║ ╚═╝ ██║██║  ██║██║ ╚████║██║  ██║╚██████╔╝███████╗██║  ██║
╚══════╝╚═╝  ╚═╝    ╚═╝     ╚═╝╚═╝  ╚═╝╚═╝  ╚═══╝╚═╝  ╚═╝ ╚═════╝ ╚══════╝╚═╝  ╚═╝
        """)
        print("=" * 80)
        print()
        print("Você gostaria de fazer qual operação?")
        print()
        print("1 - Salvar relatório de hoje")
        print("2 - Salvar relatório com intervalo específico")
        print("3 - Verificar se existem faturas a_verificar")
        print("0 - Sair")
        print()
        
        opcao = input("Escolha uma opção (0-3): ").strip()
        
        if opcao == "1":
            gerar_relatorio_hoje()
        elif opcao == "2":
            gerar_relatorio_intervalo()
        elif opcao == "3":
            verificar_faturas_pendentes()
        elif opcao == "0":
            print("\n👋 Até logo!")
            break
        else:
            print("\n❌ Opção inválida!")
            input("\nPressione ENTER para continuar...")

if __name__ == "__main__":
    # Verificar se banco existe
    if not os.path.exists('database/faturas.db'):
        print("❌ Banco de dados não encontrado!")
        print("Execute o robô primeiro: python robo.py")
        sys.exit(1)
    
    menu_principal()
